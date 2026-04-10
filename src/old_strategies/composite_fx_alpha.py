#!/usr/bin/env python
"""
Composite FX Alpha — Phase 1: Momentum + Volatility Timing

Single-script implementation following vectorbtpro conventions:
- Numba @njit kernels using vbt.generic.nb.* for rolling computations
- vbt.IF (Indicator Factory) for signal computation
- vbt.Portfolio.from_signals with signal_func_nb for simulation
- vbt accessors (.vbt) for analysis and plotting
"""

import os
from typing import Any

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import vectorbtpro as vbt
from numba import njit
from plotly.subplots import make_subplots

pd.set_option("display.max_columns", None)


# ═══════════════════════════════════════════════════════════════════════
# 1. NUMBA KERNELS
# ═══════════════════════════════════════════════════════════════════════


@njit
def momentum_signal_nb(close: np.ndarray, w_short: int, w_long: int) -> np.ndarray:
    """Blended momentum: 0.5 * log_return(21d) + 0.5 * log_return(63d)."""
    n = len(close)
    out = np.full(n, np.nan)
    for i in range(w_long, n):
        r_short = np.log(close[i] / close[i - w_short])
        r_long = np.log(close[i] / close[i - w_long])
        out[i] = 0.5 * r_short + 0.5 * r_long
    return out


@njit
def regime_weight_nb(
    vr: np.ndarray,
    low_th: float,
    high_th: float,
    w_low: float,
    w_normal: float,
    w_high: float,
) -> np.ndarray:
    """Map volatility regime ratio to momentum weight (Table 3)."""
    n = len(vr)
    out = np.full(n, np.nan)
    for i in range(n):
        if np.isnan(vr[i]):
            continue
        elif vr[i] < low_th:
            out[i] = w_low
        elif vr[i] > high_th:
            out[i] = w_high
        else:
            out[i] = w_normal
    return out


@njit
def vol_scaling_nb(ewma_vol: np.ndarray, target: float, cap: float) -> np.ndarray:
    """λ_t = min(σ* / σ_P,t, λ_max) (Eq. 13)."""
    n = len(ewma_vol)
    out = np.full(n, np.nan)
    for i in range(n):
        v = ewma_vol[i]
        if not np.isnan(v) and v > 0:
            out[i] = min(target / v, cap)
    return out


@njit
def drawdown_control_nb(
    dd: np.ndarray,
    soft: float,
    hard: float,
    recovery: float,
) -> np.ndarray:
    """State machine: NORMAL(1.0) → REDUCED(0.5) → FLAT(0.0) with hysteresis."""
    n = len(dd)
    mult = np.ones(n)
    state = 0  # 0=NORMAL, 1=REDUCED, 2=FLAT
    for i in range(n):
        d = dd[i]
        if state == 0:
            if d > hard:
                state = 2
            elif d > soft:
                state = 1
        elif state == 1:
            if d > hard:
                state = 2
            elif d < recovery:
                state = 0
        elif state == 2 and d < recovery:
            state = 0
        if state == 1:
            mult[i] = 0.5
        elif state == 2:
            mult[i] = 0.0
    return mult


@njit
def sub_portfolio_weights_nb(
    direction: np.ndarray,
    regime_wt: np.ndarray,
    vol_scale: np.ndarray,
    dd_mult: np.ndarray,
    n_days: int,
    k: int,
) -> np.ndarray:
    """K=5 overlapping sub-portfolio weights (Jegadeesh-Titman).

    Sub-portfolio j rebalances every k*5 trading days, staggered by j*5 days.
    Between rebalances, weights are held constant. Returns the average.
    """
    sub_w = np.zeros((k, n_days))
    for j in range(k):
        current = 0.0
        for i in range(n_days):
            if i >= j * 5 and (i - j * 5) % (k * 5) == 0:
                d = direction[i]
                r = regime_wt[i]
                v = vol_scale[i]
                m = dd_mult[i]
                if not (np.isnan(d) or np.isnan(r) or np.isnan(v) or np.isnan(m)):
                    current = d * r * v * m
            sub_w[j, i] = current

    weights = np.zeros(n_days)
    for i in range(n_days):
        total = 0.0
        for j in range(k):
            total += sub_w[j, i]
        weights[i] = total / k
    return weights


@njit
def compute_composite_nb(
    close: np.ndarray,
    returns: np.ndarray,
    w_short: int,
    w_long: int,
    vol_short: int,
    vol_long: int,
    ewma_span: int,
    target_vol: float,
    leverage_cap: float,
    vr_low: float,
    vr_high: float,
    mom_w_low: float,
    mom_w_normal: float,
    mom_w_high: float,
    dd_soft: float,
    dd_hard: float,
    dd_recovery: float,
    n_sub: int,
) -> tuple:
    """Master numba kernel: compute all signals → daily target weights."""
    n = len(close)

    # 1. Momentum signal + direction
    momentum = momentum_signal_nb(close, w_short, w_long)
    direction = np.full(n, 0.0)
    for i in range(n):
        if not np.isnan(momentum[i]):
            direction[i] = (
                1.0 if momentum[i] > 0 else (-1.0 if momentum[i] < 0 else 0.0)
            )

    # 2. Rolling realized vol (short + long windows) → regime ratio
    sigma_short = vbt.generic.nb.rolling_std_1d_nb(
        returns, vol_short, minp=vol_short, ddof=1
    )
    sigma_long = vbt.generic.nb.rolling_std_1d_nb(
        returns, vol_long, minp=vol_long, ddof=1
    )
    vr = np.full(n, np.nan)
    for i in range(n):
        if (
            not np.isnan(sigma_short[i])
            and not np.isnan(sigma_long[i])
            and sigma_long[i] > 0
        ):
            vr[i] = sigma_short[i] / sigma_long[i]

    # 3. Regime weight
    regime_wt = regime_weight_nb(
        vr, vr_low, vr_high, mom_w_low, mom_w_normal, mom_w_high
    )

    # 4. EWMA vol + scaling factor
    sq_returns = np.empty(n)
    for i in range(n):
        sq_returns[i] = returns[i] ** 2 if not np.isnan(returns[i]) else np.nan
    ewma_var = vbt.generic.nb.ewm_mean_1d_nb(sq_returns, ewma_span, minp=1, adjust=True)
    ewma_vol = np.full(n, np.nan)
    for i in range(n):
        if not np.isnan(ewma_var[i]) and ewma_var[i] > 0:
            ewma_vol[i] = np.sqrt(ewma_var[i]) * np.sqrt(252)
    vol_scale = vol_scaling_nb(ewma_vol, target_vol, leverage_cap)

    # 5. Proxy equity + drawdown control
    proxy_equity = np.ones(n)
    for i in range(1, n):
        d, v, r = direction[i - 1], vol_scale[i - 1], returns[i]
        if not (np.isnan(d) or np.isnan(v) or np.isnan(r)):
            proxy_equity[i] = proxy_equity[i - 1] * (1 + r * d * v)
        else:
            proxy_equity[i] = proxy_equity[i - 1]

    lookback = 63
    dd = np.zeros(n)
    for i in range(n):
        peak = proxy_equity[max(0, i - lookback + 1)]
        for j in range(max(0, i - lookback + 1), i + 1):
            if proxy_equity[j] > peak:
                peak = proxy_equity[j]
        dd[i] = 1.0 - proxy_equity[i] / peak if peak > 0 else 0.0

    dd_mult = drawdown_control_nb(dd, dd_soft, dd_hard, dd_recovery)

    # 6. K overlapping sub-portfolio weights
    weights = sub_portfolio_weights_nb(
        direction, regime_wt, vol_scale, dd_mult, n, n_sub
    )

    return momentum, direction, vr, regime_wt, ewma_vol, vol_scale, dd, dd_mult, weights


# ═══════════════════════════════════════════════════════════════════════
# 2. INDICATOR FACTORY
# ═══════════════════════════════════════════════════════════════════════

CompositeAlpha = vbt.IF(
    class_name="CompositeAlpha",
    short_name="ca",
    input_names=["close", "returns"],
    param_names=[
        "w_short",
        "w_long",
        "vol_short",
        "vol_long",
        "ewma_span",
        "target_vol",
        "leverage_cap",
        "vr_low",
        "vr_high",
        "mom_w_low",
        "mom_w_normal",
        "mom_w_high",
        "dd_soft",
        "dd_hard",
        "dd_recovery",
        "n_sub",
    ],
    output_names=[
        "momentum",
        "direction",
        "vol_regime",
        "regime_weight",
        "ewma_vol",
        "vol_scale",
        "drawdown",
        "dd_multiplier",
        "target_weight",
    ],
).with_apply_func(
    compute_composite_nb,
    takes_1d=True,
    w_short=21,
    w_long=63,
    vol_short=21,
    vol_long=252,
    ewma_span=30,
    target_vol=0.10,
    leverage_cap=3.0,
    vr_low=0.8,
    vr_high=1.2,
    mom_w_low=0.20,
    mom_w_normal=0.30,
    mom_w_high=0.50,
    dd_soft=0.12,
    dd_hard=0.20,
    dd_recovery=0.10,
    n_sub=5,
)


# ═══════════════════════════════════════════════════════════════════════
# 3. SIGNAL FUNCTION (for Portfolio.from_signals)
# ═══════════════════════════════════════════════════════════════════════


@njit
def composite_signal_nb(c, target_weights, size_arr):
    """Signal function that rebalances to target weight each bar.

    Computes delta between current position (units) and target position,
    then fires entry/exit signals with size in units (Amount).
    Handles 4 cases: add long, reduce long, add short, reduce short.
    Direction switches use upon_opposite_entry=Reverse.
    """
    tw = vbt.pf_nb.select_nb(c, target_weights)
    if np.isnan(tw):
        return False, False, False, False

    pos = c.last_position[c.col]
    price = c.last_val_price[c.col]
    value = c.last_value[c.group]

    if value <= 0 or price <= 0:
        return False, False, False, False

    target_pos = tw * value / price
    delta = target_pos - pos

    # Only rebalance if delta > 0.5% of portfolio value
    if abs(delta * price / value) < 0.005:
        return False, False, False, False

    size_arr[c.i, c.col] = abs(delta)

    if pos >= 0 and delta > 0:
        return True, False, False, False  # add to long
    elif pos > 0 and delta < 0 and target_pos >= 0:
        return False, True, False, False  # trim long
    elif pos >= 0 and target_pos < 0:
        return False, False, True, False  # flip to short (Reverse)
    elif pos <= 0 and delta < 0:
        return False, False, True, False  # add to short
    elif pos < 0 and delta > 0 and target_pos <= 0:
        return False, False, False, True  # trim short
    elif pos <= 0 and target_pos > 0:
        return True, False, False, False  # flip to long (Reverse)

    return False, False, False, False


# ═══════════════════════════════════════════════════════════════════════
# 4. PLOTTING
# ═══════════════════════════════════════════════════════════════════════


def plot_signal_dashboard(daily: pd.DataFrame, ca) -> go.Figure:
    """5-panel signal analysis dashboard."""
    idx = daily.index
    fig = make_subplots(
        rows=5,
        cols=1,
        shared_xaxes=True,
        vertical_spacing=0.04,
        subplot_titles=[
            "EUR-USD Close",
            "Momentum Signal",
            "Vol Regime (VR_t)",
            "Vol Scaling (λ_t)",
            "Target Weight (K=5 avg)",
        ],
    )
    fig.add_trace(
        go.Scatter(x=idx, y=daily["close"], name="Close", line={"width": 1}),
        row=1,
        col=1,
    )

    mom = ca.momentum.values.flatten()
    fig.add_trace(
        go.Scatter(x=idx, y=mom, name="Momentum", line={"width": 1}), row=2, col=1
    )
    fig.add_hline(y=0, line_dash="dash", line_color="gray", row=2, col=1)

    vr = ca.vol_regime.values.flatten()
    fig.add_trace(
        go.Scatter(x=idx, y=vr, name="VR_t", line={"width": 1, "color": "#636EFA"}),
        row=3,
        col=1,
    )
    fig.add_hline(y=0.8, line_dash="dot", line_color="green", row=3, col=1)
    fig.add_hline(y=1.2, line_dash="dot", line_color="red", row=3, col=1)

    vs = ca.vol_scale.values.flatten()
    fig.add_trace(
        go.Scatter(x=idx, y=vs, name="λ_t", line={"width": 1, "color": "#00CC96"}),
        row=4,
        col=1,
    )

    tw = ca.target_weight.values.flatten()
    fig.add_trace(
        go.Scatter(x=idx, y=tw, name="Weight", line={"width": 1}, fill="tozeroy"),
        row=5,
        col=1,
    )

    fig.update_layout(
        title="Composite FX Alpha — Signal Dashboard", height=1200, showlegend=False
    )
    return fig


def plot_monthly_heatmap(pf: vbt.Portfolio) -> go.Figure:
    """Monthly returns heatmap."""
    rets = pf.returns
    monthly = rets.resample("ME").apply(lambda x: (1 + x).prod() - 1)
    df = pd.DataFrame(
        {
            "year": monthly.index.year,
            "month": monthly.index.month,
            "ret": monthly.values,
        }
    )
    pivot = df.pivot(index="year", columns="month", values="ret")
    pivot.columns = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]
    fig = go.Figure(
        data=go.Heatmap(
            z=pivot.values,
            x=pivot.columns,
            y=pivot.index.astype(str),
            colorscale="RdYlGn",
            zmid=0,
            text=np.round(pivot.values * 100, 1),
            texttemplate="%{text}%",
        )
    )
    fig.update_layout(title="Monthly Returns Heatmap (%)", height=400)
    return fig


# ═══════════════════════════════════════════════════════════════════════
# 5. CROSS-VALIDATION PIPELINE
# ═══════════════════════════════════════════════════════════════════════


def optimized_pipeline(
    close,
    w_short: int = 21,
    w_long: int = 63,
    vol_short: int = 21,
    vol_long: int = 252,
    ewma_span: int = 30,
    target_vol: float = 0.10,
    leverage_cap: float = 5.0,
    vr_low: float = 0.8,
    vr_high: float = 1.2,
    mom_w_low: float = 0.20,
    mom_w_normal: float = 0.30,
    mom_w_high: float = 0.50,
    dd_soft: float = 0.12,
    dd_hard: float = 0.20,
    dd_recovery: float = 0.10,
    n_sub: int = 5,
    metric: str = "sharpe_ratio",
    fees: float = 0.00035,
    init_cash: float = 1_000_000,
    leverage: float = 2.0,
    **kwargs,
):
    """Run the full strategy pipeline and return a single performance metric.

    Designed to be wrapped by vbt.cv_split — receives a sliced close series
    per split window and scalar parameter values per grid combination.
    """
    if not isinstance(close, pd.Series):
        close = pd.Series(close)

    log_returns = np.log(close / close.shift(1))

    ca = CompositeAlpha.run(
        close=close,
        returns=log_returns,
        w_short=w_short,
        w_long=w_long,
        vol_short=vol_short,
        vol_long=vol_long,
        ewma_span=ewma_span,
        target_vol=target_vol,
        leverage_cap=leverage_cap,
        vr_low=vr_low,
        vr_high=vr_high,
        mom_w_low=mom_w_low,
        mom_w_normal=mom_w_normal,
        mom_w_high=mom_w_high,
        dd_soft=dd_soft,
        dd_hard=dd_hard,
        dd_recovery=dd_recovery,
        n_sub=n_sub,
        jitted_loop=True,
        jitted_warmup=True,
    )

    pf = vbt.Portfolio.from_signals(
        close=close,
        signal_func_nb=composite_signal_nb,
        signal_args=(
            vbt.Rep("target_weights"),
            vbt.Rep("size"),
        ),
        broadcast_named_args={
            "target_weights": ca.target_weight.values,
        },
        size=vbt.RepEval("np.full(wrapper.shape_2d, np.nan)"),
        size_type="amount",
        accumulate=True,
        upon_opposite_entry="Reverse",
        fees=fees,
        init_cash=init_cash,
        leverage=leverage,
        leverage_mode="lazy",
        freq="1D",
    )

    return pf.deep_getattr(metric)


def run_cross_validation(
    close: pd.Series,
    param_grid: dict[str, Any],
    metric: str = "sharpe_ratio",
    cv_split_kwargs: dict | None = None,
    fees: float = 0.00035,
    init_cash: float = 1_000_000,
    leverage: float = 2.0,
) -> Any:
    """Run cross-validation over a parameter grid using vbt.cv_split.

    Parameters
    ----------
    close : pd.Series
        Daily close prices (the only takeable arg — sliced by the splitter).
    param_grid : dict
        Parameters to search. Use vbt.Param([...]) for searched params,
        scalar values for fixed params.
    metric : str
        Performance metric to optimize (e.g. 'sharpe_ratio', 'total_return').
    cv_split_kwargs : dict, optional
        Override default splitter configuration.
    """
    # length=1200 (~4.8 years): 600 train + 600 test per window.
    # The strategy needs ~252 days warmup (vol_long), leaving ~348
    # usable signal days per train fold.
    default_cv_kwargs = {
        "splitter": "from_n_rolling",
        "splitter_kwargs": {
            "n": 5,
            "length": 1200,
            "split": 0.5,
            "set_labels": ["train", "test"],
        },
    }

    if cv_split_kwargs:
        default_cv_kwargs.update(cv_split_kwargs)

    cv_pipeline = vbt.cv_split(
        optimized_pipeline,
        **default_cv_kwargs,
        takeable_args=["close"],
        parameterized_kwargs={
            "engine": "threadpool",
            "chunk_len": "auto",
        },
        merge_func="concat",
    )

    for key, value in param_grid.items():
        if isinstance(value, list) and not isinstance(value, vbt.Param):
            param_grid[key] = vbt.Param(value)

    result = cv_pipeline(
        close=close,
        metric=metric,
        fees=fees,
        init_cash=init_cash,
        leverage=leverage,
        _return_grid="all",
        _index=close.index,
        **param_grid,
    )

    return result


def extract_best_parameters(
    cv_result: Any,
    param_names: list[str] | None = None,
) -> dict[str, Any]:
    """Extract best parameters from cross-validation results.

    Parameters
    ----------
    cv_result : tuple or pd.Series
        Result from run_cross_validation — (grid_perf, best_perf) tuple
        or a single Series.
    param_names : list of str, optional
        Parameter names to extract. Defaults to all 16 CompositeAlpha params.
    """
    if param_names is None:
        param_names = [
            "w_short",
            "w_long",
            "vol_short",
            "vol_long",
            "ewma_span",
            "target_vol",
            "leverage_cap",
            "vr_low",
            "vr_high",
            "mom_w_low",
            "mom_w_normal",
            "mom_w_high",
            "dd_soft",
            "dd_hard",
            "dd_recovery",
            "n_sub",
        ]

    if isinstance(cv_result, tuple) and len(cv_result) >= 2:
        best_perf = cv_result[1]
    else:
        best_perf = cv_result

    if not isinstance(best_perf, pd.Series) or not hasattr(best_perf, "index"):
        raise ValueError("Input must be a pandas Series with parameter index")

    best_params: dict[str, Any] = {}

    if isinstance(best_perf.index, pd.MultiIndex):
        level_names = best_perf.index.names
        for param in param_names:
            if param in level_names:
                level_pos = level_names.index(param)
                unique_values = best_perf.index.get_level_values(level_pos).unique()
                if len(unique_values) == 1:
                    best_params[param] = unique_values[0]
                else:
                    max_val = best_perf.max()
                    best_rows = best_perf[best_perf == max_val]
                    best_values = best_rows.index.get_level_values(level_pos)
                    best_params[param] = pd.Series(best_values).mode()[0]

    if not best_params:
        defaults = {
            "w_short": 21,
            "w_long": 63,
            "vol_short": 21,
            "vol_long": 252,
            "ewma_span": 30,
            "target_vol": 0.10,
            "leverage_cap": 5.0,
            "vr_low": 0.8,
            "vr_high": 1.2,
            "mom_w_low": 0.20,
            "mom_w_normal": 0.30,
            "mom_w_high": 0.50,
            "dd_soft": 0.12,
            "dd_hard": 0.20,
            "dd_recovery": 0.10,
            "n_sub": 5,
        }
        for param in param_names:
            best_params[param] = defaults.get(param)

    return best_params


def save_backtest_results_to_excel(
    ticker: str,
    pf: vbt.Portfolio,
    pf_opt: vbt.Portfolio,
    best_perf: pd.Series,
    standard_params: dict[str, Any],
    optimal_params: dict[str, Any],
    metric: str = "sharpe_ratio",
    file_path: str | None = None,
) -> None:
    """Save standard vs optimized backtest results to a formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    if file_path is None:
        file_path = f"results/composite_fx_alpha/{ticker}_cv_results.xlsx"
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    wb = Workbook()
    header_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    param_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    highlight_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )

    # ── Sheet 1: Backtest Results ──
    ws1 = wb.active
    ws1.title = f"{ticker} Backtest Results"
    ws1.cell(1, 1, f"Backtest Results for {ticker}").font = Font(size=16, bold=True)
    ws1.cell(2, 1, f"Optimized for: {metric}").font = Font(size=12, italic=True)

    def _write_section(ws, row, title, params, pf_obj):
        ws.cell(row, 1, title).font = header_font
        ws.cell(row, 1).fill = header_fill
        ws.merge_cells(f"A{row}:D{row}")
        row += 1
        ws.cell(row, 1, "Parameters:").font = subheader_font
        ws.cell(row, 1).fill = param_fill
        ws.merge_cells(f"A{row}:D{row}")
        row += 1
        for k, v in params.items():
            ws.cell(row, 1, k).font = normal_font
            ws.cell(row, 2, str(v)).font = normal_font
            ws.cell(row, 1).border = thin_border
            ws.cell(row, 2).border = thin_border
            row += 1
        row += 1
        ws.cell(row, 1, "Performance Metrics:").font = subheader_font
        ws.cell(row, 1).fill = header_fill
        ws.merge_cells(f"A{row}:D{row}")
        row += 1
        stats_df = pf_obj.stats().reset_index()
        stats_df.columns = ["Metric", "Value"]
        for _, r in stats_df.iterrows():
            cell_font = normal_font
            cell_fill = None
            if metric.lower() in str(r["Metric"]).lower():
                cell_font = Font(size=10, bold=True)
                cell_fill = highlight_fill
            ws.cell(row, 1, r["Metric"]).font = cell_font
            ws.cell(row, 2, str(r["Value"])).font = cell_font
            if cell_fill:
                ws.cell(row, 1).fill = cell_fill
                ws.cell(row, 2).fill = cell_fill
            ws.cell(row, 1).border = thin_border
            ws.cell(row, 2).border = thin_border
            row += 1
        return row + 2

    row = 4
    row = _write_section(ws1, row, "Standard Backtest", standard_params, pf)
    _write_section(
        ws1, row, f"Optimized Backtest (for {metric})", optimal_params, pf_opt
    )

    # ── Sheet 2: Cross-Validation Results ──
    ws2 = wb.create_sheet(title="Cross-Validation Results")
    ws2.cell(1, 1, f"Cross-Validation Results for {ticker}").font = Font(
        size=16, bold=True
    )

    cv_df = best_perf.reset_index() if isinstance(best_perf, pd.Series) else best_perf
    if cv_df.columns[-1] == 0 or cv_df.columns[-1] == "Performance":
        cv_df.rename(columns={cv_df.columns[-1]: metric}, inplace=True)

    row = 3
    for col_idx, col_name in enumerate(cv_df.columns, 1):
        ws2.cell(row, col_idx, str(col_name)).font = header_font
        ws2.cell(row, col_idx).fill = header_fill
        ws2.cell(row, col_idx).border = thin_border
    row += 1
    for df_row in dataframe_to_rows(cv_df, index=False, header=False):
        for col_idx, val in enumerate(df_row, 1):
            ws2.cell(row, col_idx, val).font = normal_font
            ws2.cell(row, col_idx).border = thin_border
        row += 1

    wb.save(file_path)
    print(f"  Excel results saved to {file_path}")


# ═══════════════════════════════════════════════════════════════════════
# 6. MAIN
# ═══════════════════════════════════════════════════════════════════════


if __name__ == "__main__":
    from utils import load_fx_data

    # ── Load & resample (VBT Pro native) ──────────────────────────
    print("Loading EUR-USD minute data via vbt.Data...")
    _, data = load_fx_data(shift_hours=7)
    data_daily = data.resample("1d")
    daily = data_daily.get()
    close = data_daily.close
    log_returns = np.log(close / close.shift(1))
    print(
        f"  {len(daily)} trading days: {daily.index[0].date()} → {daily.index[-1].date()}"
    )

    # ── A. Standard backtest ──────────────────────────────────────
    standard_params = {
        "w_short": 21,
        "w_long": 63,
        "vol_short": 21,
        "vol_long": 252,
        "ewma_span": 30,
        "target_vol": 0.10,
        "leverage_cap": 5.0,
        "vr_low": 0.8,
        "vr_high": 1.2,
        "mom_w_low": 0.20,
        "mom_w_normal": 0.30,
        "mom_w_high": 0.50,
        "dd_soft": 0.12,
        "dd_hard": 0.20,
        "dd_recovery": 0.10,
        "n_sub": 5,
    }

    print("Computing Composite Alpha signals...")
    ca = CompositeAlpha.run(
        close=close, returns=log_returns, jitted_loop=True, jitted_warmup=True
    )

    print(
        f"  Long: {int((ca.direction == 1).sum())} days, Short: {int((ca.direction == -1).sum())} days"
    )
    print(
        f"  Vol scale: [{ca.vol_scale.min():.2f}, {ca.vol_scale.max():.2f}], mean={ca.vol_scale.mean():.2f}"
    )
    print(f"  Max proxy DD: {ca.drawdown.max():.2%}")
    print(
        f"  Weight range: [{ca.target_weight.min():.3f}, {ca.target_weight.max():.3f}]"
    )

    print("Running standard backtest...")
    pf = vbt.Portfolio.from_signals(
        close=close,
        signal_func_nb=composite_signal_nb,
        signal_args=(
            vbt.Rep("target_weights"),
            vbt.Rep("size"),
        ),
        broadcast_named_args={
            "target_weights": ca.target_weight.values,
        },
        size=vbt.RepEval("np.full(wrapper.shape_2d, np.nan)"),
        size_type="amount",
        accumulate=True,
        upon_opposite_entry="Reverse",
        fees=0.00035,
        init_cash=1_000_000,
        leverage=2.0,
        leverage_mode="lazy",
        freq="1D",
    )

    print("\n" + "=" * 60)
    print("COMPOSITE FX ALPHA — STANDARD BACKTEST")
    print("=" * 60)
    print(pf.stats().to_string())
    print("=" * 60)

    # ── B. Cross-validation ───────────────────────────────────────
    optimization_metric = "sharpe_ratio"

    param_grid = {
        "w_short": vbt.Param([10, 15, 21, 30]),
        "w_long": vbt.Param([42, 63, 84, 126]),
        "ewma_span": vbt.Param([15, 30, 45, 60]),
        "target_vol": vbt.Param([0.06, 0.08, 0.10, 0.12, 0.15]),
    }

    print(f"\nRunning cross-validation ({optimization_metric})...")
    print(
        f"  Grid: {4 * 4 * 4 * 5} combinations x 5 splits = {4 * 4 * 4 * 5 * 5} backtests"
    )

    grid_perf, best_perf = run_cross_validation(
        close=close,
        param_grid=param_grid,
        metric=optimization_metric,
    )

    print(f"  Grid shape: {grid_perf.shape}")
    print(f"  Best perf shape: {best_perf.shape}")

    # ── C. Extract best & re-run ──────────────────────────────────
    best_params = extract_best_parameters(
        (grid_perf, best_perf),
        param_names=["w_short", "w_long", "ewma_span", "target_vol"],
    )
    print(f"\nBest parameters: {best_params}")

    optimal_params = {**standard_params, **best_params}

    print("Running optimized backtest...")
    ca_opt = CompositeAlpha.run(
        close=close,
        returns=log_returns,
        jitted_loop=True,
        jitted_warmup=True,
        **optimal_params,
    )

    pf_opt = vbt.Portfolio.from_signals(
        close=close,
        signal_func_nb=composite_signal_nb,
        signal_args=(
            vbt.Rep("target_weights"),
            vbt.Rep("size"),
        ),
        broadcast_named_args={
            "target_weights": ca_opt.target_weight.values,
        },
        size=vbt.RepEval("np.full(wrapper.shape_2d, np.nan)"),
        size_type="amount",
        accumulate=True,
        upon_opposite_entry="Reverse",
        fees=0.00035,
        init_cash=1_000_000,
        leverage=2.0,
        leverage_mode="lazy",
        freq="1D",
    )

    # ── D. Compare & save ─────────────────────────────────────────
    print("\n" + "=" * 60)
    print("STANDARD vs OPTIMIZED")
    print("=" * 60)
    comparison = pd.DataFrame(
        {
            "Standard": pf.stats(),
            "Optimized": pf_opt.stats(),
        }
    )
    print(comparison.to_string())
    print("=" * 60)

    results_dir = "results/composite_fx_alpha"
    os.makedirs(results_dir, exist_ok=True)

    # Portfolio overview (standard)
    fig_pf = pf.plot(
        subplots=["cumulative_returns", "drawdowns", "underwater", "trade_pnl"]
    )
    fig_pf.update_layout(title="Composite FX Alpha — Standard Portfolio", height=1000)
    fig_pf.write_html(f"{results_dir}/portfolio_overview.html")
    fig_pf.show()

    # Portfolio overview (optimized)
    fig_pf_opt = pf_opt.plot(
        subplots=["cumulative_returns", "drawdowns", "underwater", "trade_pnl"]
    )
    fig_pf_opt.update_layout(
        title="Composite FX Alpha — Optimized Portfolio", height=1000
    )
    fig_pf_opt.write_html(f"{results_dir}/portfolio_overview_optimized.html")
    fig_pf_opt.show()

    # Signal dashboards
    fig_signals = plot_signal_dashboard(daily, ca)
    fig_signals.write_html(f"{results_dir}/signal_dashboard.html")
    fig_signals.show()

    fig_signals_opt = plot_signal_dashboard(daily, ca_opt)
    fig_signals_opt.update_layout(
        title="Composite FX Alpha — Optimized Signal Dashboard"
    )
    fig_signals_opt.write_html(f"{results_dir}/signal_dashboard_optimized.html")
    fig_signals_opt.show()

    # Monthly heatmaps
    fig_monthly = plot_monthly_heatmap(pf)
    fig_monthly.write_html(f"{results_dir}/monthly_returns.html")
    fig_monthly.show()

    fig_monthly_opt = plot_monthly_heatmap(pf_opt)
    fig_monthly_opt.update_layout(title="Monthly Returns Heatmap — Optimized (%)")
    fig_monthly_opt.write_html(f"{results_dir}/monthly_returns_optimized.html")
    fig_monthly_opt.show()

    # CV heatmap
    fig_heatmap = grid_perf.vbt.heatmap(
        x_level="w_short",
        y_level="target_vol",
        slider_level="split",
    )
    fig_heatmap.write_html(f"{results_dir}/cv_heatmap.html")
    fig_heatmap.show()

    # Excel report
    save_backtest_results_to_excel(
        ticker="EUR-USD",
        pf=pf,
        pf_opt=pf_opt,
        best_perf=best_perf,
        standard_params=standard_params,
        optimal_params=optimal_params,
        metric=optimization_metric,
    )

    print(f"\nAll reports saved to {results_dir}/")
