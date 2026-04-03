#!/usr/bin/env python
"""
Trading Strategy Cross-Validation Script

This script implements an intraday momentum strategy with flexible
cross-validation functionality using vectorbtpro.
"""

import pandas as pd
import numpy as np
from typing import Tuple, Dict, List, Union, Any, Optional
import vectorbtpro as vbt
from numba import njit
import plotly.graph_objects as go

# Configure pandas display options
pd.set_option("display.max_columns", None)


# ---------------------------------------------------------------------------
# Helper functions from paste.txt (all Numba-optimized functions)
# ---------------------------------------------------------------------------


@njit
def find_day_boundaries_nb(index_ns: np.ndarray) -> Tuple[np.ndarray, np.ndarray, int]:
    """Return the start / end indices of each trading day."""
    n = len(index_ns)
    start_idx = np.empty(n, dtype=np.int64)
    end_idx = np.empty(n, dtype=np.int64)

    if n == 0:
        return start_idx, end_idx, 0

    # Convert timestamps to sequential day numbers (NumPy int)
    day_number = vbt.dt_nb.days_nb(ts=index_ns)
    current_day = day_number[0]
    day_counter = 0
    current_start = 0

    # Detect a boundary whenever ``day_number`` changes value.
    for i in range(1, n):
        if day_number[i] != current_day:
            start_idx[day_counter] = current_start
            end_idx[day_counter] = i
            day_counter += 1
            current_day = day_number[i]
            current_start = i

    # Append the final day.
    start_idx[day_counter] = current_start
    end_idx[day_counter] = n
    day_counter += 1

    return start_idx, end_idx, day_counter


@njit
def compute_abs_move_from_open_nb(
    index_ns: np.ndarray,
    close_minute: np.ndarray,
    open_minute: np.ndarray,
) -> np.ndarray:
    """Compute |Close / FirstOpen − 1| for every intraday bar."""
    n = len(index_ns)
    move_open = np.full(n, np.nan)
    if n == 0:
        return move_open

    day_number = vbt.dt_nb.days_nb(ts=index_ns)
    current_day = day_number[0]
    current_start = 0

    # Iterate once over the array, assigning the absolute move for the
    # slice corresponding to each day.
    for i in range(1, n):
        if day_number[i] != current_day:
            slice_ = slice(current_start, i)
            first_open = open_minute[current_start]
            move_open[slice_] = np.abs(close_minute[slice_] / first_open - 1.0)
            current_start = i
            current_day = day_number[i]

    # Handle last day
    if current_start < n:
        slice_ = slice(current_start, n)
        first_open = open_minute[current_start]
        move_open[slice_] = np.abs(close_minute[slice_] / first_open - 1.0)

    return move_open


@njit
def compute_sigma_open_nb(
    index_ns: np.ndarray,
    close_minute: np.ndarray,
    open_minute: np.ndarray,
    window_size: int,
) -> np.ndarray:
    """Return *sigma_open*: mean(|Close − Open|) by minute‑of‑day lagged 1 bar."""
    move_open = compute_abs_move_from_open_nb(index_ns, close_minute, open_minute)
    n = len(move_open)
    rolled = np.full(n, np.nan)

    minute_of_day = vbt.dt_nb.minute_nb(ts=index_ns)
    hour_of_day = vbt.dt_nb.hour_nb(ts=index_ns)
    group_key = hour_of_day * 60 + minute_of_day  # unique minute id

    minp = window_size - 1  # require full window for stability
    for minute_id in np.unique(group_key):
        idx = np.where(group_key == minute_id)[0]
        if idx.size:
            sub = move_open[idx]
            rolled[idx] = vbt.generic.nb.rolling_mean_1d_nb(
                sub,
                window=window_size,
                minp=minp,
            )

    sigma_open = np.full(n, np.nan)
    sigma_open[1:] = rolled[:-1]  # lag by one bar
    return sigma_open


@njit
def compute_daily_rolling_volatility_nb(
    index_ns: np.ndarray,
    close_minute: np.ndarray,
    window_size: int,
) -> np.ndarray:
    """Compute close‑to‑close rolling volatility and broadcast to minutes."""
    n = len(close_minute)
    if n == 0 or window_size <= 0:
        return np.full(n, np.nan)

    start_arr, end_arr, n_days = find_day_boundaries_nb(index_ns)
    if n_days < 2:
        return np.full(n, np.nan)

    # Extract the last close of each session
    last_close = np.full(n_days, np.nan)
    for d in range(n_days):
        last_close[d] = close_minute[end_arr[d] - 1]

    # Compute simple returns R_t = Close_t / Close_{t‑1} − 1
    returns = np.full(n_days - 1, np.nan)
    for i in range(1, n_days):
        prev = last_close[i - 1]
        if np.abs(prev) > 1e-9:
            returns[i - 1] = last_close[i] / prev - 1.0

    rolling_std = vbt.generic.nb.rolling_std_1d_nb(
        returns,
        window=window_size,
        minp=window_size,
        ddof=1,
    )

    vol_per_minute = np.full(n, np.nan)
    for d in range(1, n_days):
        std_val = rolling_std[d - 1] if d - 1 < rolling_std.size else np.nan
        vol_per_minute[start_arr[d] : end_arr[d]] = std_val

    return vol_per_minute


@njit
def compute_leverage_nb(
    rolling_vol_per_minute: np.ndarray,
    sigma_target: float,
    max_leverage: float,
    use_leverage: bool,
) -> np.ndarray:
    """Compute volatility‑targeted leverage capped at ``max_leverage``."""
    n = len(rolling_vol_per_minute)
    leverage = np.full(n, 1.0)

    if not use_leverage:
        return leverage

    for i in range(n):
        vol = rolling_vol_per_minute[i]
        if not np.isnan(vol) and vol > 0.0:
            val = sigma_target / vol
            leverage[i] = val if val <= max_leverage else max_leverage

    return leverage


@njit
def compute_intraday_bands_nb(
    start_arr: np.ndarray,
    end_arr: np.ndarray,
    n_days: int,
    close_minute: np.ndarray,
    open_minute: np.ndarray,
    sigma_open: np.ndarray,
    band_mult: float,
) -> Tuple[np.ndarray, np.ndarray]:
    """Return (upper_band, lower_band) arrays for every intraday bar."""
    n = len(close_minute)
    upper_band = np.full(n, np.nan)
    lower_band = np.full(n, np.nan)

    for d in range(1, n_days):  # first day lacks prior close
        start = start_arr[d]
        end = end_arr[d]

        previous_close = close_minute[end_arr[d - 1] - 1]
        first_open = open_minute[start]

        intraday_high_anchor = max(first_open, previous_close)
        intraday_low_anchor = min(first_open, previous_close)

        for i in range(start, end):
            sigma_val = sigma_open[i]
            upper_band[i] = intraday_high_anchor * (1.0 + band_mult * sigma_val)
            lower_band[i] = intraday_low_anchor * (1.0 - band_mult * sigma_val)

    return upper_band, lower_band


@njit
def compute_bands_nb(
    index_ns: np.ndarray,
    high_minute: np.ndarray,
    low_minute: np.ndarray,
    volume_minute: np.ndarray,
    close_minute: np.ndarray,
    open_minute: np.ndarray,
    window_size: int,
    band_mult: float,
    max_leverage: float,
    sigma_target: float,
    use_leverage: bool,
) -> Tuple[
    np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray
]:
    """Compute all intraday indicators required by the strategy."""
    n = len(close_minute)
    if n == 0:
        nan_arr = np.full(n, np.nan)
        return nan_arr, nan_arr, nan_arr, nan_arr, nan_arr, nan_arr, nan_arr

    abs_move_open = compute_abs_move_from_open_nb(index_ns, close_minute, open_minute)
    sigma_open = compute_sigma_open_nb(index_ns, close_minute, open_minute, window_size)
    rolling_vol = compute_daily_rolling_volatility_nb(
        index_ns, close_minute, window_size
    )
    leverage = compute_leverage_nb(
        rolling_vol, sigma_target, max_leverage, use_leverage
    )

    start_arr, end_arr, n_days = find_day_boundaries_nb(index_ns)
    upper_band, lower_band = compute_intraday_bands_nb(
        start_arr,
        end_arr,
        n_days,
        close_minute,
        open_minute,
        sigma_open,
        band_mult,
    )

    # Group lengths for VWAP aggregation per day
    group_lens = end_arr[:n_days] - start_arr[:n_days]
    vwap = vbt.indicators.nb.vwap_1d_nb(
        high_minute,
        low_minute,
        close_minute,
        volume_minute,
        group_lens,
    )

    return (
        upper_band,
        lower_band,
        sigma_open,
        abs_move_open,
        rolling_vol,
        leverage,
        vwap,
    )


@njit
def intraday_signal_nb(
    c,
    close_price_arr: np.ndarray,
    upper_band_arr: np.ndarray,
    lower_band_arr: np.ndarray,
    vwap_arr: np.ndarray,
    index_ns_arr: np.ndarray,
    eod_exit_trigger_hour_param: np.ndarray,
    eod_exit_trigger_minute_param: np.ndarray,
):
    """Return a 4‑tuple (entry_long, exit_long, entry_short, exit_short)."""
    # Extract current intraday time components
    ts_ns = index_ns_arr[c.i]
    cur_hour = vbt.dt_nb.hour_nb(ts_ns)
    cur_minute = vbt.dt_nb.minute_nb(ts_ns)

    # Select scalar EOD trigger values for the current column c.col
    selected_eod_hour = vbt.pf_nb.select_nb(c, eod_exit_trigger_hour_param)
    selected_eod_minute = vbt.pf_nb.select_nb(c, eod_exit_trigger_minute_param)

    # Default signals
    entry_long = False
    exit_long = False
    entry_short = False
    exit_short = False

    # 1. EOD forced exit logic
    is_eod_period = (cur_hour > selected_eod_hour) or (
        cur_hour == selected_eod_hour and cur_minute >= selected_eod_minute
    )

    if is_eod_period:
        # Force exit of open positions and prohibit new entries
        if vbt.pf_nb.ctx_helpers.in_long_position_nb(c):
            exit_long = True
        if vbt.pf_nb.ctx_helpers.in_short_position_nb(c):
            exit_short = True
        return False, exit_long, False, exit_short

    # 2. 30‑minute bar logic (on‑the‑minute evaluation)
    if cur_minute % 30 == 0:  # Act only at minutes :00 and :30
        # Pull current indicator values for the current bar and column
        close_px = vbt.pf_nb.select_nb(c, close_price_arr)
        up_band = vbt.pf_nb.select_nb(c, upper_band_arr)
        low_band = vbt.pf_nb.select_nb(c, lower_band_arr)
        vwap_val = vbt.pf_nb.select_nb(c, vwap_arr)

        # Skip if any NaNs are present (data gap)
        if (
            np.isnan(close_px)
            or np.isnan(up_band)
            or np.isnan(low_band)
            or np.isnan(vwap_val)
        ):
            return False, False, False, False

        # Entry / exit rules
        in_long_pos = vbt.pf_nb.ctx_helpers.in_long_position_nb(c)
        in_short_pos = vbt.pf_nb.ctx_helpers.in_short_position_nb(c)

        if not in_long_pos and not in_short_pos:  # If no open position
            # --- Entries ---
            if close_px > up_band:
                entry_long = True
                return entry_long, exit_long, entry_short, exit_short
            elif close_px < low_band:
                entry_short = True
                return entry_long, exit_long, entry_short, exit_short

        elif in_long_pos:
            # --- Long exit ---
            long_exit_threshold = max(up_band, vwap_val)
            if close_px < long_exit_threshold:
                exit_long = True
                return entry_long, exit_long, entry_short, exit_short

        elif in_short_pos:
            # --- Short exit ---
            short_exit_threshold = min(low_band, vwap_val)
            if close_px > short_exit_threshold:
                exit_short = True
                return entry_long, exit_long, entry_short, exit_short

    # Default – no signal generated
    return False, False, False, False


def optimized_pipeline(
    index_ns: np.ndarray,
    high: pd.Series,
    low: pd.Series,
    close: pd.Series,
    open: pd.Series,
    volume: pd.Series,
    window_size: int = 10,
    band_mult: float = 1.0,
    max_leverage: float = 4.0,
    sigma_target: float = 0.02,
    use_leverage: bool = False,
    eod_exit_trigger_hour: int = 15,
    eod_exit_trigger_minute: int = 45,
    metric: str = "sharpe_ratio",
    **kwargs,  # For portfolio parameters like fixed_fees
):
    """Run an optimized trading pipeline with intraday momentum strategy.

    Parameters
    ----------
    index_ns : np.ndarray
        Nanosecond timestamp array
    high : pd.Series
        High prices
    low : pd.Series
        Low prices
    close : pd.Series
        Close prices
    open : pd.Series
        Open prices
    volume : pd.Series
        Volume data
    window_size : int, default 10
        Lookback window size for volatility calculations
    band_mult : float, default 1.0
        Multiplier for band width
    max_leverage : float, default 4.0
        Maximum allowed leverage
    sigma_target : float, default 0.02
        Target volatility for position sizing
    use_leverage : bool, default False
        Whether to use dynamic leverage based on volatility
    eod_exit_trigger_hour : int, default 15
        Hour to trigger end-of-day exit (24h format)
    eod_exit_trigger_minute : int, default 45
        Minute to trigger end-of-day exit
    metric : str, default 'sharpe_ratio'
        Performance metric to return
    **kwargs
        Additional parameters passed to Portfolio.from_signals

    Returns
    -------
    float or pd.Series
        The selected performance metric
    """
    IMI = vbt.IF(
        class_name="IntradayMomentumIndicator",
        short_name="imi",
        input_names=[
            "index_ns",
            "high_minute",
            "low_minute",
            "volume_minute",
            "close_minute",
            "open_minute",
        ],
        param_names=[
            "window_size",
            "band_mult",
            "max_leverage",
            "sigma_target",
            "use_leverage",
        ],
        output_names=[
            "upper_band",
            "lower_band",
            "sigma_open",
            "abs_move_open",
            "rolling_vol",
            "leverage",
            "vwap",
        ],
    ).with_apply_func(
        compute_bands_nb,
        takes_1d=True,
        window_size=window_size,
        band_mult=band_mult,
        max_leverage=max_leverage,
        sigma_target=sigma_target,
        use_leverage=use_leverage,
    )

    # Instantiate the indicator
    imi = IMI.run(
        index_ns=index_ns,
        high_minute=high,
        low_minute=low,
        volume_minute=volume,
        close_minute=close,
        open_minute=open,
        window_size=window_size,
        band_mult=band_mult,
        max_leverage=max_leverage,
        sigma_target=sigma_target,
        use_leverage=use_leverage,
    )

    # Build the Portfolio with custom signal function
    pf = vbt.Portfolio.from_signals(
        close,  # Use close price series directly
        signal_func_nb=intraday_signal_nb,
        signal_args=(
            vbt.Rep("close_arr"),
            vbt.Rep("upper_band_arr"),
            vbt.Rep("lower_band_arr"),
            vbt.Rep("vwap_arr"),
            vbt.Rep("index_arr"),
            vbt.Rep("eod_exit_trigger_hour"),
            vbt.Rep("eod_exit_trigger_minute"),
        ),
        broadcast_named_args=dict(
            close_arr=close,
            upper_band_arr=imi.upper_band.values,
            lower_band_arr=imi.lower_band.values,
            vwap_arr=imi.vwap.values,
            index_arr=index_ns,
            eod_exit_trigger_hour=eod_exit_trigger_hour,
            eod_exit_trigger_minute=eod_exit_trigger_minute,
        ),
        leverage=imi.leverage.values,
        **kwargs,  # Pass kwargs (e.g., fixed_fees) to the portfolio
    )

    return pf.deep_getattr(metric)


def run_pipeline(
    minute_data: vbt.HDFData,
    index_ns: np.ndarray,
    window_size: int = 10,
    band_mult: float = 1.0,
    max_leverage: float = 4.0,
    sigma_target: float = 0.02,
    use_leverage: bool = False,
    eod_exit_trigger_hour: int = 15,
    eod_exit_trigger_minute: int = 45,
    fixed_fees: float = 0.0035,
):
    """Return (*portfolio*, *indicator*) built from intraday momentum logic."""

    # ------------------------------------------------------------------
    # 1. Build custom indicator class (vectorbt interface factory)
    # ------------------------------------------------------------------
    IMIBase = vbt.IF(
        class_name="IntradayMomentumIndicator",
        short_name="imi",
        input_names=[
            "index_ns",
            "high_minute",
            "low_minute",
            "volume_minute",
            "close_minute",
            "open_minute",
        ],
        param_names=[
            "window_size",
            "band_mult",
            "max_leverage",
            "sigma_target",
            "use_leverage",
        ],
        output_names=[
            "upper_band",
            "lower_band",
            "sigma_open",
            "abs_move_open",
            "rolling_vol",
            "leverage",
            "vwap",
        ],
    ).with_apply_func(
        compute_bands_nb,
        takes_1d=True,
        window_size=window_size,
        band_mult=band_mult,
        max_leverage=max_leverage,
        sigma_target=sigma_target,
        use_leverage=use_leverage,
    )

    class IntradayMomentumIndicator(IMIBase):  # type: ignore[misc]
        """Indicator subclass adding a convenience Plotly chart."""

        def plot(
            self,
            column: tuple | str | int | None = None,
            fig: go.Figure | None = None,
            **layout_kwargs,
        ) -> go.Figure:
            close = self.select_col_from_obj(self.close_minute, column).rename("Close")
            upper = self.select_col_from_obj(self.upper_band, column).rename(
                "Upper Band"
            )
            lower = self.select_col_from_obj(self.lower_band, column).rename(
                "Lower Band"
            )
            vwap_line = self.select_col_from_obj(self.vwap, column).rename("VWAP")

            fig = fig or go.Figure()

            close.vbt.plot(
                fig=fig,
                trace_kwargs=dict(
                    name="Close",
                    line=dict(width=2, color="blue"),
                ),
            )

            # Étape 1: Tracer la bande inférieure (ligne grise, sans remplissage "tonexty" ici)
            lower.vbt.plot(
                fig=fig,
                trace_kwargs=dict(
                    name="Lower Band",
                    line=dict(width=1.2, color="grey"),
                    # Pas de 'fill' ici si on utilise la technique de remplissage
                    # de la bande supérieure vers celle-ci.
                ),
            )

            # Étape 2: Tracer la bande supérieure (ligne grise)
            # et la remplir vers la bande inférieure ("tonexty")
            upper.vbt.plot(
                fig=fig,
                trace_kwargs=dict(
                    name="Upper Band",
                    line=dict(width=1.2, color="grey"),
                    fill="tonexty",  # Remplit vers la trace Y précédente (lower band)
                    fillcolor="rgba(255, 255, 0, 0.2)",  # Votre couleur de remplissage jaune
                ),
            )

            vwap_line.vbt.plot(
                fig=fig,
                trace_kwargs=dict(
                    name="VWAP", line=dict(color="red", width=1, dash="dot")
                ),
            )

            fig.update_layout(**layout_kwargs)
            return fig

    # Instantiate the indicator over the full minute data set.
    imi = IntradayMomentumIndicator.run(
        index_ns=index_ns,
        high_minute=minute_data.high,
        low_minute=minute_data.low,
        volume_minute=minute_data.volume,
        close_minute=minute_data.close,
        open_minute=minute_data.open,
        window_size=window_size,
        band_mult=band_mult,
        max_leverage=max_leverage,
        sigma_target=sigma_target,
        use_leverage=use_leverage,
        jitted_loop=True,
        jitted_warmup=True,
        execute_kwargs=dict(engine="threadpool", n_chunks="auto"),
    )

    # ------------------------------------------------------------------
    # 2. Build the Portfolio with custom signal function
    # ------------------------------------------------------------------
    pf = vbt.Portfolio.from_signals(
        minute_data,
        signal_func_nb=intraday_signal_nb,
        signal_args=(
            vbt.Rep("close_arr"),
            vbt.Rep("upper_band_arr"),
            vbt.Rep("lower_band_arr"),
            vbt.Rep("vwap_arr"),
            vbt.Rep("index_arr"),
            vbt.Rep("eod_exit_trigger_hour"),
            vbt.Rep("eod_exit_trigger_minute"),
        ),
        broadcast_named_args=dict(
            close_arr=minute_data.close.values,
            upper_band_arr=imi.upper_band.values,
            lower_band_arr=imi.lower_band.values,
            vwap_arr=imi.vwap.values,
            index_arr=index_ns,
            eod_exit_trigger_hour=eod_exit_trigger_hour,
            eod_exit_trigger_minute=eod_exit_trigger_minute,
        ),
        leverage=imi.leverage.values,
        fixed_fees=fixed_fees,
    )

    return pf, imi


def run_pipeline_single(
    index_ns: np.ndarray,
    high: pd.Series,
    low: pd.Series,
    close: pd.Series,
    open: pd.Series,
    volume: pd.Series,
    window_size: int = 10,
    band_mult: float = 1.0,
    max_leverage: float = 4.0,
    sigma_target: float = 0.02,
    use_leverage: bool = False,
    eod_exit_trigger_hour: int = 15,
    eod_exit_trigger_minute: int = 45,
    fixed_fees: float = 0.0035,
):
    """Run pipeline for a single asset using individual Series.

    Parameters
    ----------
    index_ns : np.ndarray
        Nanosecond timestamp array
    high : pd.Series
        High prices
    low : pd.Series
        Low prices
    close : pd.Series
        Close prices
    open : pd.Series
        Open prices
    volume : pd.Series
        Volume data
    window_size : int, default 10
        Lookback window size for volatility calculations
    band_mult : float, default 1.0
        Multiplier for band width
    max_leverage : float, default 4.0
        Maximum allowed leverage
    sigma_target : float, default 0.02
        Target volatility for position sizing
    use_leverage : bool, default False
        Whether to use dynamic leverage based on volatility
    eod_exit_trigger_hour : int, default 15
        Hour to trigger end-of-day exit (24h format)
    eod_exit_trigger_minute : int, default 45
        Minute to trigger end-of-day exit
    fixed_fees : float, default 0.0035
        Transaction cost per trade

    Returns
    -------
    Tuple[vbt.Portfolio, Any]
        Portfolio object and indicator object
    """
    # ------------------------------------------------------------------
    # 1. Build custom indicator class (vectorbt interface factory)
    # ------------------------------------------------------------------
    IMIBase = vbt.IF(
        class_name="IntradayMomentumIndicator",
        short_name="imi",
        input_names=[
            "index_ns",
            "high_minute",
            "low_minute",
            "volume_minute",
            "close_minute",
            "open_minute",
        ],
        param_names=[
            "window_size",
            "band_mult",
            "max_leverage",
            "sigma_target",
            "use_leverage",
        ],
        output_names=[
            "upper_band",
            "lower_band",
            "sigma_open",
            "abs_move_open",
            "rolling_vol",
            "leverage",
            "vwap",
        ],
    ).with_apply_func(
        compute_bands_nb,
        takes_1d=True,
        window_size=window_size,
        band_mult=band_mult,
        max_leverage=max_leverage,
        sigma_target=sigma_target,
        use_leverage=use_leverage,
    )

    class IntradayMomentumIndicator(IMIBase):  # type: ignore[misc]
        """Indicator subclass adding a convenience Plotly chart."""

        def plot(
            self,
            column: tuple | str | int | None = None,
            fig: go.Figure | None = None,
            **layout_kwargs,
        ) -> go.Figure:
            close = self.select_col_from_obj(self.close_minute, column).rename("Close")
            upper = self.select_col_from_obj(self.upper_band, column).rename(
                "Upper Band"
            )
            lower = self.select_col_from_obj(self.lower_band, column).rename(
                "Lower Band"
            )
            vwap_line = self.select_col_from_obj(self.vwap, column).rename("VWAP")

            fig = fig or go.Figure()

            close.vbt.plot(
                fig=fig,
                trace_kwargs=dict(
                    name="Close",
                    line=dict(width=2, color="blue"),
                ),
            )

            # Étape 1: Tracer la bande inférieure (ligne grise, sans remplissage "tonexty" ici)
            lower.vbt.plot(
                fig=fig,
                trace_kwargs=dict(
                    name="Lower Band",
                    line=dict(width=1.2, color="grey"),
                    # Pas de 'fill' ici si on utilise la technique de remplissage
                    # de la bande supérieure vers celle-ci.
                ),
            )

            # Étape 2: Tracer la bande supérieure (ligne grise)
            # et la remplir vers la bande inférieure ("tonexty")
            upper.vbt.plot(
                fig=fig,
                trace_kwargs=dict(
                    name="Upper Band",
                    line=dict(width=1.2, color="grey"),
                    fill="tonexty",  # Remplit vers la trace Y précédente (lower band)
                    fillcolor="rgba(255, 255, 0, 0.2)",  # Votre couleur de remplissage jaune
                ),
            )

            vwap_line.vbt.plot(
                fig=fig,
                trace_kwargs=dict(
                    name="VWAP", line=dict(color="red", width=1, dash="dot")
                ),
            )

            fig.update_layout(**layout_kwargs)
            return fig

    # Instantiate the indicator over the full minute data set.
    imi = IntradayMomentumIndicator.run(
        index_ns=index_ns,
        high_minute=high,
        low_minute=low,
        volume_minute=volume,
        close_minute=close,
        open_minute=open,
        window_size=window_size,
        band_mult=band_mult,
        max_leverage=max_leverage,
        sigma_target=sigma_target,
        use_leverage=use_leverage,
        jitted_loop=True,
        jitted_warmup=True,
        execute_kwargs=dict(engine="threadpool", n_chunks="auto"),
    )

    # ------------------------------------------------------------------
    # 2. Build the Portfolio with custom signal function
    # ------------------------------------------------------------------
    pf = vbt.Portfolio.from_signals(
        close,  # Use close price series directly
        signal_func_nb=intraday_signal_nb,
        signal_args=(
            vbt.Rep("close_arr"),
            vbt.Rep("upper_band_arr"),
            vbt.Rep("lower_band_arr"),
            vbt.Rep("vwap_arr"),
            vbt.Rep("index_arr"),
            vbt.Rep("eod_exit_trigger_hour"),
            vbt.Rep("eod_exit_trigger_minute"),
        ),
        broadcast_named_args=dict(
            close_arr=close.values,
            upper_band_arr=imi.upper_band.values,
            lower_band_arr=imi.lower_band.values,
            vwap_arr=imi.vwap.values,
            index_arr=index_ns,
            eod_exit_trigger_hour=eod_exit_trigger_hour,
            eod_exit_trigger_minute=eod_exit_trigger_minute,
        ),
        leverage=imi.leverage.values,
        fixed_fees=fixed_fees,
    )

    return pf, imi


def run_backtest(
    minute_df: vbt.HDFData,
    ticker: str,
    window_size: int = 10,
    band_mult: float = 1.0,
    max_leverage: float = 4.0,
    sigma_target: float = 0.02,
    use_leverage: bool = False,
    eod_exit_trigger_hour: int = 15,
    eod_exit_trigger_minute: int = 45,
    fixed_fees: float = 0.0035,
) -> Tuple[vbt.Portfolio, Any]:
    """Run a backtest for a single set of parameters.

    Parameters
    ----------
    minute_df : vbt.HDFData
        HDF data container with market data for multiple tickers
    ticker : str
        Ticker symbol to use for the strategy (e.g., 'TSLA')
    window_size : int, default 10
        Lookback window size for volatility calculations
    band_mult : float, default 1.0
        Multiplier for band width
    max_leverage : float, default 4.0
        Maximum allowed leverage
    sigma_target : float, default 0.02
        Target volatility for position sizing
    use_leverage : bool, default False
        Whether to use dynamic leverage based on volatility
    eod_exit_trigger_hour : int, default 15
        Hour to trigger end-of-day exit (24h format)
    eod_exit_trigger_minute : int, default 45
        Minute to trigger end-of-day exit
    fixed_fees : float, default 0.0035
        Transaction cost per trade
    **kwargs
        Additional parameters passed to the pipeline

    Returns
    -------
    Tuple[vbt.Portfolio, Any]
        Portfolio object and indicator object
    """
    # Convert index to nanoseconds
    idx_ns = vbt.dt.to_ns(minute_df.index)

    # Extract data series for the ticker
    high = minute_df.data[ticker]["High"]
    low = minute_df.data[ticker]["Low"]
    close = minute_df.data[ticker]["Close"]
    open_price = minute_df.data[ticker]["Open"]
    volume = minute_df.data[ticker]["Volume"]

    # Run the pipeline for a single asset
    pf, imi = run_pipeline_single(
        index_ns=idx_ns,
        high=high,
        low=low,
        close=close,
        open=open_price,
        volume=volume,
        window_size=window_size,
        band_mult=band_mult,
        max_leverage=max_leverage,
        sigma_target=sigma_target,
        use_leverage=use_leverage,
        eod_exit_trigger_hour=eod_exit_trigger_hour,
        eod_exit_trigger_minute=eod_exit_trigger_minute,
        fixed_fees=fixed_fees,
    )

    return pf, imi


def run_cross_validation(
    minute_df: vbt.HDFData,
    ticker: str,
    param_grid: Dict[str, Union[List, vbt.Param]],
    metric: str = "total_return",
    cv_split_kwargs: Optional[Dict] = None,
    fixed_fees: float = 0.0035,
    attach_bounds: str = "index",
    **additional_kwargs,
) -> Union[Any, Tuple[Any, Any]]:
    """Run cross-validation for the trading strategy with flexible parameters.

    Parameters
    ----------
    minute_df : vbt.HDFData
        HDF data container with market data for multiple tickers
    ticker : str
        Ticker symbol to use for the strategy (e.g., 'TSLA')
    param_grid : Dict[str, Union[List, vbt.Param]]
        Dictionary of parameters to optimize, can include:
        - window_size
        - band_mult
        - max_leverage
        - sigma_target
        - use_leverage
        - eod_exit_trigger_hour
        - eod_exit_trigger_minute
    metric : str, default 'total_return'
        Performance metric to optimize
    cv_split_kwargs : Dict, optional
        Additional kwargs for the cv_split function
    fixed_fees : float, default 0.0035
        Transaction cost per trade
    attach_bounds : str, default 'index'
        How to attach bounds, options: 'index', 'columns', None
    return_grid : str, default 'both'
        What to return, options: 'all', 'best', 'both'
    **additional_kwargs
        Additional parameters passed to the pipeline

    Returns
    -------
    Union[Any, Tuple[Any, Any]]
        Depending on return_grid parameter:
        - 'both': (grid_performance, best_performance)
        - 'all': grid_performance only
        - 'best': best_performance only
    """
    # Convert index to nanoseconds for timestamp calculations
    idx_ns = vbt.dt.to_ns(minute_df.index)

    # Default CV split parameters
    default_cv_kwargs = {
        "splitter": "from_ranges",
        "splitter_kwargs": dict(
            every="M",  # Monthly split
            split=0.5,  # 50% train, 50% test
            set_labels=["train", "test"],
        ),
    }

    # Update with user-provided kwargs if any
    if cv_split_kwargs:
        default_cv_kwargs.update(cv_split_kwargs)

    # Create the cross-validation wrapper for the pipeline
    cv_pipeline = vbt.cv_split(
        optimized_pipeline,
        **default_cv_kwargs,
        takeable_args=["index_ns", "high", "low", "close", "open", "volume"],
        parameterized_kwargs=dict(
            engine="threadpool",
            chunk_len="auto",
        ),
        merge_func="concat",
    )

    # Extract data for the specified ticker from the HDF container
    high = minute_df.data[ticker]["High"]
    low = minute_df.data[ticker]["Low"]
    close = minute_df.data[ticker]["Close"]
    open_price = minute_df.data[ticker]["Open"]
    volume = minute_df.data[ticker]["Volume"]

    # Create a wrapper for proper data structure
    wrapper = minute_df.symbol_wrapper

    # Convert standard parameter lists to vbt.Param objects if needed
    for key, value in param_grid.items():
        if isinstance(value, list) and not isinstance(value, vbt.Param):
            param_grid[key] = vbt.Param(value)

    # Run the cross-validation
    result = cv_pipeline(
        index_ns=idx_ns,
        high=high,
        low=low,
        close=close,
        open=open_price,
        volume=volume,
        metric=metric,
        fixed_fees=fixed_fees,
        attach_bounds=attach_bounds,
        _merge_kwargs=dict(wrapper=wrapper),
        _return_grid="all",  # Pass the return_grid parameter directly
        _index=minute_df.index,
        **param_grid,
        **additional_kwargs,
    )

    return result


def extract_best_parameters(
    cv_result: Union[pd.Series, Tuple[pd.Series, pd.Series]],
    param_names: List[str] = None,
) -> Dict[str, Any]:
    """Extract best parameters from cross-validation results.

    Parameters
    ----------
    cv_result : Union[pd.Series, Tuple[pd.Series, pd.Series]]
        Result from run_cross_validation - either a single Series (best_perf)
        or a tuple (grid_perf, best_perf)
    param_names : List[str], optional
        Names of parameters to extract. If None, will try to extract all common
        parameters (window_size, band_mult, etc.)

    Returns
    -------
    Dict[str, Any]
        Dictionary of best parameters
    """
    # Default parameter names to look for if not specified
    if param_names is None:
        param_names = [
            "window_size",
            "band_mult",
            "max_leverage",
            "sigma_target",
            "use_leverage",
            "eod_exit_trigger_hour",
            "eod_exit_trigger_minute",
        ]

    # If input is a tuple, extract the best_perf component
    if isinstance(cv_result, tuple) and len(cv_result) >= 2:
        best_perf = cv_result[1]
    else:
        best_perf = cv_result

    # Make sure we're working with a Series that has a MultiIndex
    if not isinstance(best_perf, pd.Series) or not hasattr(best_perf, "index"):
        raise ValueError(
            "Input must be a pandas Series with an index containing parameter values"
        )

    # Extract parameters from the MultiIndex
    best_params = {}

    # Check if the result is a MultiIndex Series
    if isinstance(best_perf.index, pd.MultiIndex):
        # Get the level names of the MultiIndex
        level_names = best_perf.index.names

        # Find the levels that correspond to our parameters
        for param in param_names:
            if param in level_names:
                # Get the level position
                level_pos = level_names.index(param)

                # Extract unique values for this parameter
                unique_values = best_perf.index.get_level_values(level_pos).unique()

                # If there's only one value, use it
                if len(unique_values) == 1:
                    best_params[param] = unique_values[0]
                # Otherwise, find the value that appears in the best-performing rows
                else:
                    # Get the values from rows with the highest performance
                    max_val = best_perf.max()
                    best_rows = best_perf[best_perf == max_val]
                    best_values = best_rows.index.get_level_values(level_pos)

                    # Use the most common value
                    best_params[param] = pd.Series(best_values).mode()[0]

    # If no parameters were found, provide defaults
    if not best_params:
        defaults = {
            "window_size": 10,
            "band_mult": 1.0,
            "max_leverage": 4.0,
            "sigma_target": 0.02,
            "use_leverage": False,
            "eod_exit_trigger_hour": 15,
            "eod_exit_trigger_minute": 45,
        }

        for param in param_names:
            best_params[param] = defaults.get(param)

    return best_params


def save_backtest_results_to_excel(
    ticker: str,
    pf: vbt.Portfolio,
    pf_opt: vbt.Portfolio,
    best_perf: pd.Series,
    standard_params: Dict[str, Any],
    optimal_params: Dict[str, Any],
    metric: str = "total_return",
    file_path: str = None,
):
    """
    Save backtest results to a nicely formatted Excel file.

    Parameters
    ----------
    ticker : str
        The ticker symbol
    pf : vbt.Portfolio
        Standard backtest portfolio
    pf_opt : vbt.Portfolio
        Optimized backtest portfolio
    best_perf : pd.Series
        Cross-validation best performance results
    standard_params : Dict[str, Any]
        Parameters used for the standard backtest
    optimal_params : Dict[str, Any]
        Optimal parameters determined from cross-validation
    metric : str, default 'total_return'
        The performance metric that was optimized
    file_path : str, optional
        Path to save the Excel file. If None, will use '../data/results/{ticker}_backtest_results.xlsx'
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import os

    # Set default file path if not provided
    if file_path is None:
        file_path = f"../data/results/{ticker}_backtest_results.xlsx"

    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Create a new workbook
    wb = Workbook()

    # ---- Sheet 1: Backtest Results ----
    ws1 = wb.active
    ws1.title = f"{ticker} Backtest Results"

    # Add ticker as title
    ws1.cell(1, 1, f"Backtest Results for {ticker}").font = Font(size=16, bold=True)
    ws1.cell(2, 1, f"Optimized for: {metric}").font = Font(size=12, italic=True)

    # Style definitions
    header_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True, size=11)
    normal_font = Font(size=10)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    param_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )

    # ---- Standard Backtest Section ----
    # Add section header
    row = 4
    ws1.cell(row, 1, "Standard Backtest").font = header_font
    ws1.cell(row, 1).alignment = Alignment(horizontal="left")
    ws1.cell(row, 1).fill = header_fill
    ws1.merge_cells(f"A{row}:D{row}")

    # Add parameters
    row += 1
    ws1.cell(row, 1, "Parameters:").font = subheader_font
    ws1.cell(row, 1).fill = param_fill
    ws1.merge_cells(f"A{row}:D{row}")

    row += 1
    params_df = pd.DataFrame(
        {
            "Parameter": list(standard_params.keys()),
            "Value": list(standard_params.values()),
        }
    )

    for r_idx, (_, r) in enumerate(params_df.iterrows(), row):
        ws1.cell(r_idx, 1, r["Parameter"]).font = normal_font
        ws1.cell(r_idx, 2, str(r["Value"])).font = normal_font
        ws1.cell(r_idx, 1).border = thin_border
        ws1.cell(r_idx, 2).border = thin_border

    row += len(params_df) + 1

    # Add stats
    ws1.cell(row, 1, "Performance Metrics:").font = subheader_font
    ws1.cell(row, 1).fill = header_fill
    ws1.merge_cells(f"A{row}:D{row}")

    row += 1
    stats_df = pf.stats().reset_index()
    stats_df.columns = ["Metric", "Value"]

    # Highlight the optimized metric in the stats
    for r_idx, (_, r) in enumerate(stats_df.iterrows(), row):
        cell_font = normal_font
        cell_fill = None

        # If this is the metric that was optimized, highlight it
        if metric.lower() in r["Metric"].lower():
            cell_font = Font(size=10, bold=True)
            cell_fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )

        ws1.cell(r_idx, 1, r["Metric"]).font = cell_font
        ws1.cell(r_idx, 2, str(r["Value"])).font = cell_font

        if cell_fill:
            ws1.cell(r_idx, 1).fill = cell_fill
            ws1.cell(r_idx, 2).fill = cell_fill

        ws1.cell(r_idx, 1).border = thin_border
        ws1.cell(r_idx, 2).border = thin_border

    row += len(stats_df) + 2

    # ---- Optimized Backtest Section ----
    # Add section header
    ws1.cell(row, 1, f"Optimized Backtest (for {metric})").font = header_font
    ws1.cell(row, 1).alignment = Alignment(horizontal="left")
    ws1.cell(row, 1).fill = header_fill
    ws1.merge_cells(f"A{row}:D{row}")

    # Add parameters
    row += 1
    ws1.cell(row, 1, "Optimal Parameters:").font = subheader_font
    ws1.cell(row, 1).fill = param_fill
    ws1.merge_cells(f"A{row}:D{row}")

    row += 1
    opt_params_df = pd.DataFrame(
        {
            "Parameter": list(optimal_params.keys()),
            "Value": list(optimal_params.values()),
        }
    )

    for r_idx, (_, r) in enumerate(opt_params_df.iterrows(), row):
        ws1.cell(r_idx, 1, r["Parameter"]).font = normal_font
        ws1.cell(r_idx, 2, str(r["Value"])).font = normal_font
        ws1.cell(r_idx, 1).border = thin_border
        ws1.cell(r_idx, 2).border = thin_border

    row += len(opt_params_df) + 1

    # Add stats
    ws1.cell(row, 1, "Performance Metrics:").font = subheader_font
    ws1.cell(row, 1).fill = header_fill
    ws1.merge_cells(f"A{row}:D{row}")

    row += 1
    opt_stats_df = pf_opt.stats().reset_index()
    opt_stats_df.columns = ["Metric", "Value"]

    # Highlight the optimized metric in the stats
    for r_idx, (_, r) in enumerate(opt_stats_df.iterrows(), row):
        cell_font = normal_font
        cell_fill = None

        # If this is the metric that was optimized, highlight it
        if metric.lower() in r["Metric"].lower():
            cell_font = Font(size=10, bold=True)
            cell_fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )

        ws1.cell(r_idx, 1, r["Metric"]).font = cell_font
        ws1.cell(r_idx, 2, str(r["Value"])).font = cell_font

        if cell_fill:
            ws1.cell(r_idx, 1).fill = cell_fill
            ws1.cell(r_idx, 2).fill = cell_fill

        ws1.cell(r_idx, 1).border = thin_border
        ws1.cell(r_idx, 2).border = thin_border

    # Set column width
    for col in range(1, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 30

    # ---- Sheet 2: Cross-Validation Results ----
    ws2 = wb.create_sheet(title="Cross-Validation Results")

    # Add title
    ws2.cell(1, 1, f"Cross-Validation Results for {ticker}").font = Font(
        size=16, bold=True
    )
    ws2.cell(2, 1, f"Optimized for: {metric}").font = Font(size=12, italic=True)

    # Prepare the data - convert best_perf to DataFrame for easier handling
    if isinstance(best_perf.index, pd.MultiIndex):
        # Reset the multi-index to get a nice tabular format
        cv_results_df = best_perf.reset_index()
    else:
        # If it's not a MultiIndex, just convert to DataFrame
        cv_results_df = best_perf.to_frame(name="Performance").reset_index()

    # Rename the last column to the metric name
    if cv_results_df.columns[-1] == 0 or cv_results_df.columns[-1] == "Performance":
        cv_results_df.rename(columns={cv_results_df.columns[-1]: metric}, inplace=True)

    # Add data
    row = 4
    for r_idx, col_name in enumerate(cv_results_df.columns, 1):
        ws2.cell(row, r_idx, str(col_name)).font = header_font
        ws2.cell(row, r_idx).fill = header_fill
        ws2.cell(row, r_idx).border = thin_border

    for df_row in dataframe_to_rows(cv_results_df, index=False, header=False):
        row += 1
        for col_idx, val in enumerate(df_row, 1):
            ws2.cell(row, col_idx, str(val)).font = normal_font
            ws2.cell(row, col_idx).border = thin_border

    # Adjust column widths
    for col_idx, col_name in enumerate(cv_results_df.columns, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 25

    # Save the workbook
    wb.save(file_path)
    print(f"Results saved to {file_path}")


if __name__ == "__main__":
    # Live example using local HDF5 minute data (adjust path as needed)
    minute_df = vbt.HDFData.pull("../data/raw/tickers_cleaned.h5")
    idx_ns = vbt.dt.to_ns(minute_df.index)

    # Set the optimization metric
    optimization_metric = "sharpe_ratio"

    # Example 1: Run a simple backtest

    ticker = "TSLA"

    print(f"Running backtest for {ticker} with default parameters...")

    # Define standard parameters
    standard_params = {
        "window_size": 14,
        "band_mult": 1.0,
        "max_leverage": 1.0,
        "sigma_target": 0.02,
        "use_leverage": True,
        "eod_exit_trigger_hour": 15,
        "eod_exit_trigger_minute": 59,
        "fixed_fees": 0.0035,
    }

    pf, imi = run_backtest(minute_df=minute_df, ticker=ticker, **standard_params)

    print(pf.stats())
    print(imi.wrapper.columns)
    date_slice = slice("2024-04-01", "2024-04-30")
    fig = imi.xloc[date_slice].plot()
    pf.xloc[date_slice].plot_trade_signals(fig=fig).show()

    # Example 2: Run cross-validation to find optimal parameters
    print("\nRunning cross-validation to find optimal parameters...")
    param_grid = {
        "window_size": vbt.Param(
            list(range(5, 30, 1))
        ),  # Test window sizes from 10 to 45
        "band_mult": vbt.Param(
            np.linspace(start=0.5, stop=2, num=10)
        ),  # Test different band multipliers
        "max_leverage": 4.0,
        "use_leverage": False,
        "eod_exit_trigger_hour": 15,
        "eod_exit_trigger_minute": 45,
    }

    # Run cross-validation
    grid_perf, best_perf = run_cross_validation(
        minute_df=minute_df,
        ticker=ticker,
        param_grid=param_grid,
        metric=optimization_metric,
        fixed_fees=0.0035,
    )

    print(f"Grid Perf is: {grid_perf}")
    print(f"Best Perf is: {best_perf}")

    # Extract the best parameters
    best_params = extract_best_parameters((grid_perf, best_perf))
    print(f"Best Parameters: {best_params}")

    # Run backtest with the best parameters
    print("\nRunning backtest with optimal parameters...")

    # Create optimal parameters dictionary
    optimal_params = {
        "window_size": best_params.get("window_size", 10),
        "band_mult": best_params.get("band_mult", 1.0),
        "max_leverage": 4.0,
        "sigma_target": 0.02,
        "use_leverage": False,
        "eod_exit_trigger_hour": 15,
        "eod_exit_trigger_minute": 45,
        "fixed_fees": 0.0035,
    }

    # Run the optimized backtest
    pf_opt, imi_opt = run_backtest(minute_df=minute_df, ticker=ticker, **optimal_params)

    print(pf_opt.stats())

    fig_heatmap = grid_perf.vbt.heatmap(
        x_level="window_size", y_level="band_mult", slider_level="split"
    )

    fig_heatmap.show()

    # Save all results to Excel with the new path
    save_backtest_results_to_excel(
        ticker=ticker,
        pf=pf,
        pf_opt=pf_opt,
        best_perf=best_perf,
        standard_params=standard_params,
        optimal_params=optimal_params,
        metric=optimization_metric,
        file_path=f"../data/results/{ticker}_backtest_results.xlsx",
    )
